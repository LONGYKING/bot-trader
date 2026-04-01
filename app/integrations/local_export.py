"""
Local Excel export for backtest results.

Writes a multi-sheet .xlsx file to the configured backtest_export_dir.
Always called before cloud export attempts so a local copy exists regardless.

Sheets:
  - Summary  : key/value metrics
  - Trades   : one row per trade
  - Equity Curve : portfolio value over time
"""
from __future__ import annotations

import asyncio
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

import structlog

if TYPE_CHECKING:
    from app.core.backtest_engine import BacktestResult

log = structlog.get_logger(__name__)


def _write_excel(result: BacktestResult, strategy_name: str, asset: str, timeframe: str, path: Path) -> None:
    """Synchronous Excel writer — run via run_in_executor."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True)
    rows = [
        ("Metric", "Value"),
        ("Strategy", strategy_name),
        ("Asset", asset),
        ("Timeframe", timeframe),
        ("Period", f"{result.date_from[:10]} to {result.date_to[:10]}"),
        ("Initial Capital", f"${result.initial_capital:,.2f}"),
        ("Total Trades", result.total_trades),
        ("Winning Trades", result.winning_trades),
        ("Win Rate", f"{result.win_rate * 100:.1f}%"),
        ("Total PnL %", f"{result.total_pnl_pct:.2f}%"),
        ("Sharpe Ratio", f"{result.sharpe_ratio:.2f}"),
        ("Max Drawdown %", f"{result.max_drawdown_pct:.2f}%"),
        ("Annual Return %", f"{result.annual_return_pct:.2f}%"),
        ("Generated At", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, row in enumerate(rows, start=1):
        ws_summary.cell(row=i, column=1, value=row[0])
        ws_summary.cell(row=i, column=2, value=row[1])
        if i == 1:
            ws_summary.cell(row=1, column=1).font = header_font
            ws_summary.cell(row=1, column=2).font = header_font

    # --- Trades sheet ---
    ws_trades = wb.create_sheet("Trades")
    trade_headers = [
        "Entry Time", "Exit Time", "Direction", "Tenor (days)",
        "Entry Price", "Exit Price", "PnL %",
        "Capital Before ($)", "Capital After ($)", "Premium Paid ($)",
        "Trade Size ($)", "Max Exposure ($)", "Regime",
    ]
    for col, h in enumerate(trade_headers, start=1):
        cell = ws_trades.cell(row=1, column=col, value=h)
        cell.font = header_font
    for row_idx, t in enumerate(result.trades, start=2):
        ws_trades.cell(row=row_idx, column=1, value=str(t.entry_time)[:19])
        ws_trades.cell(row=row_idx, column=2, value=str(t.exit_time)[:19])
        ws_trades.cell(row=row_idx, column=3, value=t.direction)
        ws_trades.cell(row=row_idx, column=4, value=t.tenor_days)
        ws_trades.cell(row=row_idx, column=5, value=float(t.entry_price))
        ws_trades.cell(row=row_idx, column=6, value=float(t.exit_price))
        ws_trades.cell(row=row_idx, column=7, value=float(t.pnl_pct))
        ws_trades.cell(row=row_idx, column=8, value=float(t.capital_before))
        ws_trades.cell(row=row_idx, column=9, value=float(t.capital_after))
        ws_trades.cell(row=row_idx, column=10, value=float(t.premium_paid))
        ws_trades.cell(row=row_idx, column=11, value=float(t.trade_size))
        ws_trades.cell(row=row_idx, column=12, value=float(t.max_exposure))
        ws_trades.cell(row=row_idx, column=13, value=t.regime_at_entry or "")

    # --- Equity Curve sheet ---
    ws_equity = wb.create_sheet("Equity Curve")
    ws_equity.cell(row=1, column=1, value="Time").font = header_font
    ws_equity.cell(row=1, column=2, value="Portfolio Value").font = header_font
    for row_idx, point in enumerate(result.equity_curve, start=2):
        ws_equity.cell(row=row_idx, column=1, value=str(point["time"])[:19])
        ws_equity.cell(row=row_idx, column=2, value=float(point["value"]))

    wb.save(path)


async def export_backtest_to_local(
    result: BacktestResult,
    strategy_name: str,
    asset: str,
    timeframe: str,
    backtest_id: str,
    export_dir: str = "backtest_exports",
) -> Path:
    """
    Write backtest result to a local .xlsx file.
    Creates export_dir if it doesn't exist.
    Returns the Path of the written file.
    """
    out_dir = Path(export_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_name = strategy_name.replace(" ", "_").replace("/", "-")[:40]
    filename = f"{backtest_id}_{safe_name}_{asset.replace('/', '')}_{timeframe}.xlsx"
    path = out_dir / filename

    await asyncio.get_event_loop().run_in_executor(
        None, _write_excel, result, strategy_name, asset, timeframe, path
    )

    log.info("backtest_local_export_saved", path=str(path))
    return path
